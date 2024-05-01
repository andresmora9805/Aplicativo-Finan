# -*- coding: utf-8 -*-
"""
Created on Fri Mar  8 09:44:11 2024

@author: ASUS
"""

"""Costos por MW"""

#from Ejemplos import *
import numpy_financial as npf
import math 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import matplotlib.ticker as ticker
from Regulacion_Tension import calculoCalibre



def format_y_ticks(value, pos):
    # Formatear el valor con separadores de miles
    return '{:,.0f}'.format(value)


def reporteExcel(LCOE,C_InstaladaDC, energiaPrimerAno, CostoTotalInstalado, totalEgresos, valorUnitario, totalIngresos,
                 valorPorcenCostosDuros, Inversion, valorEquityTIR, plazoDeuda, impuestos, TIR,TIR_dos, VAN_dosAux,
                 VAN, DSCR, DCSRrequerido, valorDCSRpromedio, BC, impuestoBaseGravable, tarifacostoPPAlista,
                 ingresosDespuesImpuestos, interesDeudaList, contador, valorReporteEx, varLC, varCLC, track, valorGastosOyM,
                 valorAdministracionOyM, valorArriendoHa, valorAporteCliente, valorRegalias, valorPeriodoFinanciacion,
                 valorTasaInteres, valorPlazoDeuda, valorInteresDeuda, valorNumeroServicioDeuda, valorNumeroDeMesesGastoOyM,
                 valorInteresCuentaReserva, valorImpuestoRenta, depreciacion, valorTasaDepreciacionAc, valorTiempoDepreciacionLn,
                 valorTasaDepreciacionLn, valorTarifaEnergiaRed, valorCostoEnergiaPPA, valorCostoEnergiaPPAgen,
                 valorCERE, valorDuracionTarifaCostos, valorPorcentajeIncrementoTarifa, valorTasaTarifaCostos, valorCambioEntrada):
      
    letra=''
    valor=''
    contadorAux=0
    
    if valorReporteEx==1:
        book=load_workbook('Reporte Excel.xlsx')
        sheet= book.active
        maximo_filas=sheet.max_column
        contadorAux=len(book.sheetnames)
        HojaInput=book['Datos de entrada']
        if maximo_filas==3:
            letra='D'
        elif maximo_filas==4:
            letra='E'
        elif maximo_filas==5:
            letra='F'
        elif maximo_filas==6:
            letra='G'
        elif maximo_filas==7:
            letra='H'
        elif maximo_filas==8:
            letra='I'
        elif maximo_filas==9:
            letra='J'
        elif maximo_filas==10:
            letra='K'
        elif maximo_filas==11:
            letra='L'
        elif maximo_filas==12:
            letra='M'
    else:
        if contador==1:
            letra='C'
            book=Workbook()
            sheet=book.active
            HojaInput=book.create_sheet(title='Datos de entrada') 
        elif contador==2:
            letra='D'
            book=load_workbook('Reporte Excel.xlsx')
            sheet = book.active
            HojaInput=book['Datos de entrada']
        elif contador==3:
            letra='E'
            book=load_workbook('Reporte Excel.xlsx')
            sheet = book.active
            HojaInput=book['Datos de entrada']
        elif contador==4:
            letra='F'
            book=load_workbook('Reporte Excel.xlsx')
            sheet = book.active
            HojaInput=book['Datos de entrada']
        elif contador==5:
            letra='G'
            book=load_workbook('Reporte Excel.xlsx')
            sheet = book.active
            HojaInput=book['Datos de entrada']
        elif contador==6:
            letra='H'
            HojaInput=book['Datos de entrada']
        elif contador==7:
            letra='I'
            HojaInput=book['Datos de entrada']
        elif contador==8:
            letra='J'
            HojaInput=book['Datos de entrada']
        elif contador==9:
            letra='K'
            HojaInput=book['Datos de entrada']
        elif contador==10:
            letra='L'
            HojaInput=book['Datos de entrada']
        else:
            letra='M'
            HojaInput=book['Datos de entrada']
       
        
    
    sheet.title='Resumen Proyecto'
    sheet['A1'] = "Costo nivelado de energía" 
    sheet['A1'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A1'].font=Font(italic=True, bold=True)
    sheet.column_dimensions['A'].width=44
    sheet.column_dimensions['B'].width=12
    sheet.column_dimensions[letra].width=17
    sheet['B1'] = "$/kWh"
    sheet['B1'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B1'].font=Font(italic=True, bold=True)
    sheet.column_dimensions[letra].auto_size=True
    sheet[letra+'1'] = LCOE
    sheet[letra+'1'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A2'] = "Tipo de tecnología"
    sheet['A2'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A2'].font=Font(italic=True, bold=True)
    sheet['B2'] = "Fotovoltaica"
    sheet['B2'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B2'].font=Font(italic=True, bold=True)
    sheet["A3"] = "Capacidad del generador"
    sheet['A3'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A3'].font=Font(italic=True, bold=True)
    sheet["B3"] = "kWp"
    sheet['B3'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B3'].font=Font(italic=True, bold=True)
    sheet[letra+"3"] = C_InstaladaDC
    sheet[letra+'3'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A4"] = "Producción de energía"
    sheet['A4'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A4'].font=Font(italic=True, bold=True)
    sheet["B4"] = "kWh-año"
    sheet['B4'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B4'].font=Font(italic=True, bold=True)
    sheet[letra+"4"] = energiaPrimerAno
    sheet[letra+'4'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A5"] = "Vida útil proyecto"
    sheet['A5'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A5'].font=Font(italic=True, bold=True)
    sheet["B5"] = "años"
    sheet['B5'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B5'].font=Font(italic=True, bold=True)
    sheet[letra+"5"] = 25
    sheet[letra+'5'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A6"] = "Costo total instalación"
    sheet['A6'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A6'].font=Font(italic=True, bold=True)
    sheet["B6"] = "$/watt"
    sheet['B6'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B6'].font=Font(italic=True, bold=True)
    if contador==1 and valorReporteEx==0:
        sheet[letra+'6'].style=NamedStyle(name='contable', number_format='_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-')
    else:
        sheet[letra+'6'].number_format='#,##0.00'
    sheet[letra+"6"] = CostoTotalInstalado
    sheet[letra+'6'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A7"] = "Gastos operativos"
    sheet['A7'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A7'].font=Font(italic=True, bold=True)
    sheet["B7"] ="$"
    sheet['B7'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B7'].font=Font(italic=True, bold=True)
    sheet[letra+'7'].number_format='#,##0.00'
    sheet[letra+"7"] = totalEgresos[0]
    sheet[letra+'7'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A8"] = "Valor unitario de costos operacionales"
    sheet['A8'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A8'].font=Font(italic=True, bold=True)
    sheet["B8"] = "$/kWh"
    sheet['B8'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B8'].font=Font(italic=True, bold=True)
    sheet[letra+'8'].number_format='#,##0.00'
    sheet[letra+"8"] = valorUnitario
    sheet[letra+'8'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A9"] = "Total Ingresos"
    sheet['A9'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A9'].font=Font(italic=True, bold=True)
    sheet["B9"] = "$"
    sheet['B9'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B9'].font=Font(italic=True, bold=True)
    sheet[letra+'9'].number_format='#,##0.00'
    sheet[letra+"9"] = totalIngresos[0]
    sheet[letra+'9'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A10"] = "Porcentaje de costos duros"
    sheet['A10'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A10'].font=Font(italic=True, bold=True)
    sheet["B10"] = "%"
    sheet['B10'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B10'].font=Font(italic=True, bold=True)
    sheet[letra+"10"] = 100-valorPorcenCostosDuros
    sheet[letra+'10'].alignment=Alignment(horizontal='center', vertical='center')
    sheet["A11"] = "Inversión de capital"
    sheet['A11'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A11'].font=Font(italic=True, bold=True)
    sheet["B11"] = "$"
    sheet['B11'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B11'].font=Font(italic=True, bold=True)
    sheet[letra+'11'].number_format='#,##0.00'
    sheet[letra+"11"] = Inversion
    sheet[letra+'11'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A12'] = "Tir después de impuestos"
    sheet['A12'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A12'].font=Font(italic=True, bold=True)
    sheet['B12'] = "%"
    sheet['B12'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B12'].font=Font(italic=True, bold=True)
    sheet[letra+'12'] = valorEquityTIR
    sheet[letra+'12'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A13'] = "% de los costos duros financiados mediante deuda"
    sheet['A13'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A13'].font=Font(italic=True, bold=True)
    sheet['B13'] = "%"
    sheet['B13'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B13'].font=Font(italic=True, bold=True)
    sheet[letra+'13'] = valorPorcenCostosDuros
    sheet[letra+'13'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A14'] = "Plazo de la deuda"
    sheet['A14'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A14'].font=Font(italic=True, bold=True)
    sheet['B14'] = "Años"
    sheet['B14'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B14'].font=Font(italic=True, bold=True)
    sheet[letra+'14'] = plazoDeuda
    sheet[letra+'14'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A15'] ="Es el propietario sujeto a impuestos"
    sheet['A15'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A15'].font=Font(italic=True, bold=True)
    sheet[letra+'15'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'36'].alignment=Alignment(horizontal='center', vertical='center')
    if impuestos==0:
        sheet[letra+'15']= "Sí"
        HojaInput[letra+'36']="Sí"
    else:
        sheet[letra+'15']='No'
        HojaInput[letra+'36']='No'
    sheet['A16'] = "TIR antes de impuestos"
    sheet['A16'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A16'].font=Font(italic=True, bold=True)
    sheet['B16'] = "%"
    sheet['B16'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B16'].font=Font(italic=True, bold=True)
    sheet[letra+'16'] = TIR*100
    sheet[letra+'16'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A17'] = "TIR después de impuestos"
    sheet['A17'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A17'].font=Font(italic=True, bold=True)
    sheet['B17'] = "%"
    sheet['B17'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B17'].font=Font(italic=True, bold=True)
    sheet[letra+'17'] = TIR_dos*100
    sheet[letra+'17'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A18'] ="Valor presente neto de los costos"
    sheet['A18'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A18'].font=Font(italic=True, bold=True)
    sheet['B18'] = "$"
    sheet['B18'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B18'].font=Font(italic=True, bold=True)
    sheet[letra+'18'].number_format='#,##0.00'
    sheet[letra+'18'] = VAN_dosAux
    sheet[letra+'18'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A19'] = "Valor presente neto del capital final"
    sheet['A19'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A19'].font=Font(italic=True, bold=True)
    sheet['B19'] = "$"
    sheet['B19'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B19'].font=Font(italic=True, bold=True)
    sheet[letra+'19'].number_format='#,##0.00'
    sheet[letra+'19'] = VAN
    sheet[letra+'19'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A20'] = "Se cumple el DCSR mínimo requerido"
    sheet['A20'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A20'].font=Font(italic=True, bold=True)
    sheet[letra+'20'].alignment=Alignment(horizontal='center', vertical='center')
    if DSCR[0]>DCSRrequerido:
        sheet[letra+'20']="Sí"
        
    else:
        sheet[letra+'35']='No'
        
    sheet['A21']="Se cumple el DCSR promedio requerido"
    sheet['A21'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A21'].font=Font(italic=True, bold=True)
    sheet[letra+'21'].alignment=Alignment(horizontal='center', vertical='center')
    if (sum(DSCR[0:15])/len(DSCR))>valorDCSRpromedio:
        sheet[letra+'21']='Sí'
    else:
        sheet[letra+'21']='No'
        
    sheet['A22']="Período de recuperación de la inversión"
    sheet['A22'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A22'].font=Font(italic=True, bold=True)
    sheet['B22']="Años"
    sheet['B22'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B22'].font=Font(italic=True, bold=True)
    
    sheet['A23']="Relación beneficio costo"
    sheet['A23'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A23'].font=Font(italic=True, bold=True)
    sheet[letra+'23']=BC
    sheet[letra+'23'].alignment=Alignment(horizontal='center', vertical='center')
    
    sheet['A24']="Cambio realizado para comparación"
    sheet['A24'].alignment=Alignment(horizontal='center', vertical='center')
    sheet['A24'].font=Font(italic=True, bold=True)
    sheet[letra+'24']=valorCambioEntrada
    sheet[letra+'24'].alignment=Alignment(horizontal='center', vertical='center')
    
    
    HojaInput.column_dimensions['A'].width=44
    HojaInput.column_dimensions[letra].width=15
    HojaInput['A1']="Capacidad del generador"
    HojaInput['A1'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A1'].font=Font(italic=True, bold=True)
    HojaInput['B1']="kWp"
    HojaInput['B1'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['B1'].font=Font(italic=True, bold=True)
    HojaInput[letra+'1']=C_InstaladaDC
    HojaInput[letra+'1'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A2']="Distancia del predio al punto de conexión"
    HojaInput['A2'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A2'].font=Font(italic=True, bold=True)
    HojaInput[letra+'2']=varLC
    HojaInput[letra+'2'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A3']="Calibre de línea seleccionado"
    HojaInput['A3'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A3'].font=Font(italic=True, bold=True)
    HojaInput[letra+'3']=varCLC
    HojaInput[letra+'3'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A4']="¿Va a usar trackers?"
    HojaInput['A4'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A4'].font=Font(italic=True, bold=True)
    if track==1: 
        HojaInput[letra+'4']='Sí'
        HojaInput[letra+'4'].alignment=Alignment(horizontal='center', vertical='center')
    else:
        HojaInput[letra+'4']='No'
        HojaInput[letra+'4'].alignment=Alignment(horizontal='center', vertical='center')
        
    HojaInput['A6']="Variables de operación y mantenimiento"
    HojaInput['A6'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A6'].font=Font(italic=True, bold=True)
    HojaInput['A6'].fill=PatternFill(start_color="19F9EC", end_color="19F9EC", fill_type="solid")
    
    HojaInput['A7']="Costo fijo de operación y mantenimiento $/kW"
    HojaInput['A7'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A7'].font=Font(italic=True, bold=True)
    HojaInput[letra+'7']=valorGastosOyM
    HojaInput[letra+'7'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'7'].number_format='#,##0.00'
    
    HojaInput['A8']="Costos de administración $"
    HojaInput['A8'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A8'].font=Font(italic=True, bold=True)
    HojaInput[letra+'8']=valorAdministracionOyM
    HojaInput[letra+'8'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'8'].number_format='#,##0.00'
    
    HojaInput['A9']="Valor arriendo Ha $"
    HojaInput['A9'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A9'].font=Font(italic=True, bold=True)
    HojaInput[letra+'9']=valorArriendoHa
    HojaInput[letra+'9'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'9'].number_format='#,##0.00'
    
    
    HojaInput['A10']="Valor aporte cliente $"
    HojaInput['A10'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A10'].font=Font(italic=True, bold=True)
    HojaInput[letra+'10']=valorAporteCliente
    HojaInput[letra+'10'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'10'].number_format='#,##0.00'
    
    HojaInput['A11']="Valor regalias $"
    HojaInput['A11'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A11'].font=Font(italic=True, bold=True)
    HojaInput[letra+'11']=valorRegalias
    HojaInput[letra+'11'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'11'].number_format='#,##0.00'
    
    HojaInput['A20']="Variables de financiación del proyecto"
    HojaInput['A20'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A20'].font=Font(italic=True, bold=True)
    HojaInput['A20'].fill=PatternFill(start_color="19F9EC", end_color="19F9EC", fill_type="solid")    
    
    HojaInput['A21']="Financiación de la construcción"
    HojaInput['A21'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A21'].font=Font(italic=True, bold=True)
    HojaInput['A21'].fill=PatternFill(start_color="F3259F", end_color="F3259F", fill_type="solid") 
    HojaInput['A22']="Período de construcción (meses)"
    HojaInput['A22'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A22'].font=Font(italic=True, bold=True)
    HojaInput[letra+'22']=valorPeriodoFinanciacion
    HojaInput[letra+'22'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A23']="Tasa de interés efectivo anual (%)"
    HojaInput['A23'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A23'].font=Font(italic=True, bold=True)
    HojaInput[letra+'23']=valorTasaInteres
    HojaInput[letra+'23'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A24']="Financiación de la construcción a largo plazo"
    HojaInput['A24'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A24'].font=Font(italic=True, bold=True)
    HojaInput['A24'].fill=PatternFill(start_color="F3259F", end_color="F3259F", fill_type="solid") 
    HojaInput['A25']="Plazo de la deuda (años)"
    HojaInput['A25'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A25'].font=Font(italic=True, bold=True)
    HojaInput[letra+'25']=valorPlazoDeuda
    HojaInput[letra+'25'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A26']="Tasa de interés de la deuda a largo plazo (%)"
    HojaInput['A26'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A26'].font=Font(italic=True, bold=True)
    HojaInput[letra+'26']=valorInteresDeuda
    HojaInput[letra+'26'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A27']="DCSR mínimo requerido"
    HojaInput['A27'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A27'].font=Font(italic=True, bold=True)
    HojaInput[letra+'27']=DCSRrequerido
    HojaInput[letra+'27'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A28']="DCSR promedio requerido"
    HojaInput['A28'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A28'].font=Font(italic=True, bold=True)
    HojaInput[letra+'28']=valorDCSRpromedio
    HojaInput[letra+'28'].alignment=Alignment(horizontal='center', vertical='center')
    

    HojaInput['A29']="Financiación cuentas de reserva"
    HojaInput['A29'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A29'].font=Font(italic=True, bold=True)
    HojaInput['A29'].fill=PatternFill(start_color="F3259F", end_color="F3259F", fill_type="solid")
    HojaInput['A30']="Número de meses de servicio de la deuda"
    HojaInput['A30'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A30'].font=Font(italic=True, bold=True)
    HojaInput[letra+'30']=valorNumeroServicioDeuda
    HojaInput[letra+'30'].alignment=Alignment(horizontal='center', vertical='center')

    HojaInput['A31']="Reserva en operación y mantenimiento"
    HojaInput['A31'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A31'].font=Font(italic=True, bold=True)
    HojaInput['A31'].fill=PatternFill(start_color="F3259F", end_color="F3259F", fill_type="solid")
    HojaInput['A32']="Número de meses de gasto en O&M reserva (meses)"
    HojaInput['A32'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A32'].font=Font(italic=True, bold=True)
    HojaInput[letra+'32']=valorNumeroDeMesesGastoOyM
    HojaInput[letra+'32'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A33']="Intereses cuenta de reserva(%)"
    HojaInput['A33'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A33'].font=Font(italic=True, bold=True)
    HojaInput[letra+'33']=valorInteresCuentaReserva
    HojaInput[letra+'33'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A35']="Variables de impuestos y depreciación"
    HojaInput['A35'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A35'].font=Font(italic=True, bold=True)
    HojaInput['A35'].fill=PatternFill(start_color="19F9EC", end_color="19F9EC", fill_type="solid") 
    
    HojaInput['A36']="El propietario declara impuestos"
    HojaInput['A36'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A36'].font=Font(italic=True, bold=True)
        
    HojaInput['A37']="Impuesto de renta"
    HojaInput['A37'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A37'].font=Font(italic=True, bold=True)
    HojaInput[letra+'37']=valorImpuestoRenta
    HojaInput[letra+'37'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A38']="Tipo de depreciación usada"
    HojaInput['A38'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A38'].font=Font(italic=True, bold=True)
    HojaInput[letra+'38'].alignment=Alignment(horizontal='center', vertical='center')
    if depreciacion==0:
        HojaInput[letra+'38']="Lineal"
    else:
        HojaInput[letra+'38']="Acelerada"
    
    HojaInput['A39']="Tasa de depreciacion acelerada"
    HojaInput['A39'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A39'].font=Font(italic=True, bold=True)
    HojaInput[letra+'39']=valorTasaDepreciacionAc
    HojaInput[letra+'39'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A40']="Años para depreciacion lineal"
    HojaInput['A40'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A40'].font=Font(italic=True, bold=True)
    HojaInput[letra+'40']=valorTiempoDepreciacionLn
    HojaInput[letra+'40'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A41']="Tasa de depreciacion lineal"
    HojaInput['A41'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A41'].font=Font(italic=True, bold=True)
    HojaInput[letra+'41']=valorTasaDepreciacionLn
    HojaInput[letra+'41'].alignment=Alignment(horizontal='center', vertical='center')
    
    HojaInput['A43']="Variables de los costos de energía"
    HojaInput['A43'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A43'].font=Font(italic=True, bold=True)
    HojaInput['A43'].fill=PatternFill(start_color="19F9EC", end_color="19F9EC", fill_type="solid") 
    
    HojaInput['A44']="Tarifa Energía de la Red (COP/kWh)"
    HojaInput['A44'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A44'].font=Font(italic=True, bold=True)
    HojaInput[letra+'44']=valorTarifaEnergiaRed
    HojaInput[letra+'44'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'44'].number_format='#,##0.00'
    
    HojaInput['A45']="Costo energía PPA (COP/kWh)"
    HojaInput['A45'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A45'].font=Font(italic=True, bold=True)
    HojaInput[letra+'45']=valorCostoEnergiaPPA
    HojaInput[letra+'45'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'45'].number_format='#,##0.00'
    
    HojaInput['A46']="Costo energía PPA con autogenerador (COP/kWh)"
    HojaInput['A46'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A46'].font=Font(italic=True, bold=True)
    HojaInput[letra+'46']=valorCostoEnergiaPPAgen
    HojaInput[letra+'46'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'46'].number_format='#,##0.00'
    
    HojaInput['A47']="Precio CERE (COP/kWh)"
    HojaInput['A47'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A47'].font=Font(italic=True, bold=True)
    HojaInput[letra+'47']=valorCERE
    HojaInput[letra+'47'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput[letra+'47'].number_format='#,##0.00'
    
    HojaInput['A48']="Estructura tarifaría basada en costos"
    HojaInput['A48'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A48'].font=Font(italic=True, bold=True)
    HojaInput['A48'].fill=PatternFill(start_color="F3259F", end_color="F3259F", fill_type="solid")
    HojaInput['A49']="Duración de la tarifa basada en costos (años)"
    HojaInput['A49'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A49'].font=Font(italic=True, bold=True)
    HojaInput[letra+'49']=valorDuracionTarifaCostos
    HojaInput[letra+'49'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A50']="Porcentaje de incremento de la tárifa de energía (%)"
    HojaInput['A50'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A50'].font=Font(italic=True, bold=True)
    HojaInput[letra+'50']=valorPorcentajeIncrementoTarifa
    HojaInput[letra+'50'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A51']="Tasa de la tarifa basada en costos (%)"
    HojaInput['A51'].alignment=Alignment(horizontal='center', vertical='center')
    HojaInput['A51'].font=Font(italic=True, bold=True)
    HojaInput[letra+'51']=valorTasaTarifaCostos
    HojaInput[letra+'51'].alignment=Alignment(horizontal='center', vertical='center')
    
    
    
    if valorReporteEx==0:
        segundaHoja=book.create_sheet(title="Flujo de caja anual escenario"+" "+str(contador))
    else:
        segundaHoja=book.create_sheet(title="Flujo de caja anual escenario"+" "+str(contadorAux-contador-1))
    segundaHoja.column_dimensions['A'].width=17
    segundaHoja['A1']="Año del proyecto"
    segundaHoja['A1'].font=Font(italic=True,bold=True)
    segundaHoja['A1'].alignment=Alignment(horizontal='center', vertical='center')
    tiempo=[]
    GraficaLista_dos=[]
    GraficaLista_dos.append(0)
    
    for i in range(2,28):
        segundaHoja['A'+str(i)]=i-2
        segundaHoja['A'+str(i)].alignment=Alignment(horizontal='center',vertical='center')
        tiempo.append(i-2)
        if i<27:
           GraficaLista_dos.append(totalIngresos[i-2]-impuestoBaseGravable[i-2])
    
    segundaHoja.column_dimensions['B'].width=33
    segundaHoja['B1']="Tarifa Energía Vendida PPA COP/kWh"
    segundaHoja['B1'].font=Font(italic=True,bold=True)
    segundaHoja['B1'].alignment=Alignment(horizontal='center', vertical='center')
    
    segundaHoja.column_dimensions['C'].width=20
    segundaHoja['C1']="Ingresos por ventas $"
    segundaHoja['C1'].font=Font(italic=True,bold=True)
    segundaHoja['C1'].alignment=Alignment(horizontal='center', vertical='center')
    
    segundaHoja.column_dimensions['D'].width=20
    segundaHoja['D1']="Costos operativos $"
    segundaHoja['D1'].font=Font(italic=True,bold=True)
    segundaHoja['D1'].alignment=Alignment(horizontal='center', vertical='center')
    
    segundaHoja.column_dimensions['E'].width=33
    segundaHoja['E1']="Flujo de caja después de impuestos $"
    segundaHoja['E1'].font=Font(italic=True,bold=True)
    segundaHoja['E1'].alignment=Alignment(horizontal='center', vertical='center')
    
    segundaHoja.column_dimensions['F'].width=25
    segundaHoja['F1']="Flujo de caja acumulado $"
    segundaHoja['F1'].font=Font(italic=True,bold=True)
    segundaHoja['F1'].alignment=Alignment(horizontal='center', vertical='center')
    segundaHoja['F2'].alignment=Alignment(horizontal='center', vertical='center')
    segundaHoja['F2']=-Inversion
    segundaHoja['F2'].number_format='#,##0.00'
    AuxFilaF=-Inversion
    
    segundaHoja.column_dimensions['G'].width=25
    segundaHoja['G1']="Flujo de caja del interés $"
    segundaHoja['G1'].font=Font(italic=True,bold=True)
    segundaHoja['G1'].alignment=Alignment(horizontal='center', vertical='center')
    segundaHoja['G2'].alignment=Alignment(horizontal='center', vertical='center')
    AuxFilaG=0
    GraficaLista=[]
    GraficaLista.append(-Inversion)
    
    
    segundaHoja.column_dimensions['H'].width=25
    segundaHoja['H1']="Cobertura de servicio de la deuda"
    segundaHoja['H1'].font=Font(italic=True,bold=True)
    segundaHoja['H1'].alignment=Alignment(horizontal='center', vertical='center')
    
    for j in range(0,7):
        for i in range(2,27):
            if j==0:
               segundaHoja['B'+str(i+1)]=tarifacostoPPAlista[i-2]
               segundaHoja['B'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
            elif j==1:
                segundaHoja['C'+str(i+1)]=totalIngresos[i-2]
                segundaHoja['C'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                segundaHoja['C'+str(i+1)].number_format='#,##0.00'
            elif j==2:
                segundaHoja['D'+str(i+1)]=totalEgresos[i-2]
                segundaHoja['D'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                segundaHoja['D'+str(i+1)].number_format='#,##0.00'
            elif j==3:
                segundaHoja['E'+str(i+1)]=ingresosDespuesImpuestos[i-2]
                segundaHoja['E'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                segundaHoja['E'+str(i+1)].number_format='#,##0.00'
            elif j==4:
                AuxFilaF+=ingresosDespuesImpuestos[i-2]
                if AuxFilaF>0:
                    segundaHoja['F'+str(i+1)]=AuxFilaF
                    segundaHoja['F'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                    segundaHoja['F'+str(i+1)].number_format='#,##0.00'
                    GraficaLista.append(AuxFilaF)
                else:
                    sheet[letra+'22']=i
                    sheet[letra+'22'].alignment=Alignment(horizontal='center', vertical='center')
                    segundaHoja['F'+str(i+1)]=AuxFilaF
                    segundaHoja['F'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                    segundaHoja['F'+str(i+1)].number_format='#,##0.00'
                    segundaHoja['F'+str(i+1)].font=Font(color='FF0000')
                    GraficaLista.append(AuxFilaF)
            elif j==5:
                AuxFilaG+=interesDeudaList[i-2]
                segundaHoja['G'+str(i+1)]=AuxFilaG
                segundaHoja['G'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
                segundaHoja['G'+str(i+1)].number_format='#,##0.00'
            elif j==6:
                segundaHoja['H'+str(i+1)]=DSCR[i-2]
                segundaHoja['H'+str(i+1)].alignment=Alignment(horizontal='center',vertical='center')
            
    if valorReporteEx==0:
        tercerHoja=book.create_sheet(title="Graficos escenario"+" "+str(contador))
    else:
        tercerHoja=book.create_sheet(title="Graficos escenario"+" "+str(contadorAux-contador-1))
    fig, axs=plt.subplots(2, figsize=(12,6))
    axs[0].plot(tiempo,GraficaLista)
    #axs[0].set_xlabel("Años del proyecto")
    axs[0].set_ylabel("Flujo de caja acumulado")
    axs[0].set_title("Flujo de caja acumulado")
    axs[0].yaxis.set_major_formatter(ticker.FuncFormatter(format_y_ticks))
    axs[0].axhline(0, color='black', linestyle='--')
    axs[1].plot(tiempo,GraficaLista_dos)
    axs[1].set_xlabel("Años del proyecto")
    axs[1].set_ylabel("$")
    axs[1].set_title("Ingreso más beneficio en impuesto")
    axs[1].yaxis.set_major_formatter(ticker.FuncFormatter(format_y_ticks))
    axs[1].axhline(0, color='black', linestyle='--')
    fig.savefig('Grafico_1.png')
    #fig.savefig('Grafico_2.png')
    img_uno=Image('Grafico_1.png')
    #img_dos=Image('Grafico_2.png')
        
    tercerHoja.add_image(img_uno,'B1')
    #tercerHoja.add_image(img_dos,'F1')
    #print(GraficaLista_dos)
    
    
    
    
    book.save('Reporte Excel.xlsx')
    



def generarResultados(valorCapacidad, valorIrradiancia, valorAutoconsumo, valorDemanda, valorENFICC,
                          valorTarifaEnergiaRed, valorCostoEnergiaPPA, valorCostoEnergiaPPAgen, valorCERE,
                          valorArriendoHa, valorAdministracionOyM, valorTarifaPrestamista, valorPorcenCostosDuros,
                          var, track, varLC, varCLC, valorPeriodoFinanciacion, valorTasaInteres, valorPlazoDeuda,
                          valorInteresDeuda, valorDCSRminimo, valorSeguroOyM, valorPorcentajeIncrementoTarifa,valorTasaTarifaCostos,
                          valorRegalias, valorPeriodoInflacionOyM, valorGastosOyM, valorAporteCliente, valorGastosVarOyM,
                          valorInflacionOyM, valorInflacionOyM_dos, valorNumeroServicioDeuda,valorNumeroDeMesesGastoOyM,
                          valorInteresCuentaReserva, valorReemplazoEquipo, valorCostoReemplazoEquipo,valorReemplazoEquipo_dos,
                          valorCostoReemplazoEquipo_dos, depreciacion, valorTasaDepreciacionLn, valorTiempoDepreciacionLn,
                          valorTasaDepreciacionAc, impuestos, valorBaseGravable, valorMaxDesCapex, valorImpuestoRenta,
                          valorEquityTIR, contador, valorDCSRpromedio, valorCambioEntrada, valorReporteEx, valorDuracionTarifaCostos,
                          valorCostoPaneles, valorCostoInversores, valorCostoCableadoDC, valorCostoCableadoAC,
                          valorCostoSistemaMonitoreo, valorCostoTuberia, valorCostoCerramiento, valorCostoSPT, valorCostoOtros,
                          valorCostoMaquinaria, valorCostoInstalacion, valorCostoImprevistos, valorCostoIngenieria, valorCostoEstSuelos,
                          valorCostoResistividad, valorCostoEstTopografico, valorCostoPruebas, valorCostoCampamento, valorCostoSPDA,
                          valorCostoTransformador, valorCostoProtecciones, valorCostoMedidores, calculoExactoVar, valorEnergia, valorLCDig,
                          distanciaDig):
        Valor_trackers=0
        valor_Terreno=0
        Equip_Gen_Total=0
        BOS_Total=0
        Valor_conexion=0
        Total_Interconexion=0
        Capex=0
        C_InstaladaMW=valorCapacidad
        C_InstaladakW=C_InstaladaMW*1000
        C_InstaladaDC=C_InstaladakW*1.2
        Area_Proyecto=C_InstaladaMW*2
        TasaDolar=3965
        tarifaEnergiaLista=[]
        tarifacostoPPAlista=[]
        CERElista=[]
        costoPPAAGlista=[]
        vidaUtil=25
        Area_Proyecto=C_InstaladaMW*2
        Irradiancia=valorIrradiancia
        
        if calculoExactoVar==1:
            autoconsumoUsuario=valorEnergia[2]
            energiaPrimerAno=valorEnergia[3]
        else:
            autoconsumoUsuario=valorAutoconsumo
            energiaPrimerAno=C_InstaladaDC*Irradiancia*365
        demandaUsuarioSinAG=valorDemanda
        hIndisponibilidad=76.2
        degradacionEnergia=1
        listaEnergia=[]
        listaEnergia.append(energiaPrimerAno)
        ENFICC=valorENFICC/100
        ENFICCkW=0
        ENFICCkWlista=[]
        ENFICCkWlista.append(energiaPrimerAno*ENFICC)
        perdidasIndis=[]
        perdidasIndis.append(energiaPrimerAno*hIndisponibilidad/365/24)
        TasaCTFE=1
        tarifaEnergiaLista.append(valorTarifaEnergiaRed)
        costoPPA=valorCostoEnergiaPPA
        costoPPAAG=valorCostoEnergiaPPAgen
        tarifacostoPPAlista.append(costoPPA)
        CERE=valorCERE
        CERElista.append(CERE)
        costoPPAAGlista.append(costoPPAAG)
        ElectrPagadaUsuario=[]
        ElectrPagadaUsuario.append(valorTarifaEnergiaRed*demandaUsuarioSinAG)
        ElectrPagadaUsuarioAG=[]
        ElectrPagadaUsuarioAG.append(valorTarifaEnergiaRed*autoconsumoUsuario)
        regaliasTotal=[]
        costosFijos=[]
        costosVariables=[]
        gerenciayAdministracion=[]
        arrTierra=[]
        seguroList=[]
        totalEgresos=[]
        Periodo=1
        depreciacionlnAcumulada=[]
        depreciacionVal=[]
        depreciacionAcelAcumulada=[]
        valorLibro=[]
        montoDepreciar=[]
        depreciacionAux2=0
        depreciacionAux=0
        RentaLiquida=[]
        valorArriendoTotal=valorArriendoHa*Area_Proyecto
        administracion=valorAdministracionOyM
        
        """El costo de legalización por MW es de 50000 dolares"""
        
        costoMWRTB=TasaDolar*50000
        costoLegalizacion=costoMWRTB*C_InstaladaMW
        
        """Reservas y costo de financiamiento"""
        
        ReservaCostoFinan=0
        ReservaCostoFinan=valorTarifaPrestamista*valorPorcenCostosDuros
        
        if var=='Parcialmente plano':
            valor_Terreno=50000000
        elif var=='Con pendiente superior al 10%':
            valor_Terreno=65000000
        elif var=='Con vegetación':
            valor_Terreno=75000000
        elif var=='No es uniforme presenta protuberancias':
            valor_Terreno=100000000
        else:
            valor_Terreno=50000000
        
        if track==1:
            Valor_trackers=850000000 
        else:
            Valor_trackers=0
        
        Equipos_Generacion={
            "Paneles_Solares":valorCostoPaneles,
            "Inversores":valorCostoInversores,
            }
        
        BOS={
             "Cableado_DC":valorCostoCableadoDC,
             "Cableado_AC_BT":valorCostoCableadoAC,
             "Sistema_Monitoreo":valorCostoSistemaMonitoreo,
             "Trackers": Valor_trackers,
             "Tuberias_conducciones": valorCostoTuberia,
             "Cerramiento":valorCostoCerramiento,
             "SPT": valorCostoSPT,
             "Otros_materiales":valorCostoOtros,
             "Alquiler_maquinaria": valorCostoMaquinaria,
             "Adecuacion_terrenos": valor_Terreno,
             "Instalación": valorCostoInstalacion,
             "Imprevistos": valorCostoImprevistos,
             "Ingenieria_detalle": valorCostoIngenieria,
             "Estudio_suelos": valorCostoEstSuelos,
             "Estudio_topograficos": valorCostoEstTopografico,
             "Estudio_resistividad": valorCostoResistividad,
             "Pruebas_campo": valorCostoPruebas,
             "Campamento": valorCostoCampamento,
             "SPDA": valorCostoSPDA,
             }
        
        #Costos_Linea_Conexión
        Costos_LC={
            "Poste 12m-1050kg retencion": 7242621,
            "Poste 12m-510kg suspension":5509945,
            "ACSR 4 AWG": 18797225,
            "ACSR 2 AWG":21160588,
            "ACSR 1/0 AWG":22936110,
            "ACSR 2/0 AWG": 25388593,
            "ACSR 3/0 AWG": 28530033,
            "ACSR 4/0 AWG": 37373366,
            "ACSR 266 kcmil": 50239045,
            "ACSR 336 kcmil": 59884448,
            "ACSR 397 kcmil": 68948864,
            "ACSR 477 kcmil": 81187283,
            "ACSR 605 kcmil": 95211507,
            "ACSR 795 kcmil": 124576172,
            "Puesta a tierra": 462733
            }
        
        #Costo total línea conexión
        if valorLCDig==1:
            CantidadPostes=((int(distanciaDig)/25)+1)
            CantidadPostesSuspension=CantidadPostes*0.75
            CantidadPostesSuspension=round(CantidadPostesSuspension,0)
            CantidadPostesRetencion=-CantidadPostesSuspension+CantidadPostes
            Valor_conexion=Costos_LC['Poste 12m-1050kg retencion']*CantidadPostesRetencion+Costos_LC['Poste 12m-510kg suspension']*CantidadPostesSuspension+(int(distanciaDig)/1000)*Costos_LC[varCLC]+(CantidadPostes/3)*Costos_LC['Puesta a tierra']
        else:
            Metros=varLC
            MetrosAux=Metros.split('-')
            MetrosAux2=MetrosAux[1]
            MetrosAux3=MetrosAux2.replace('m','')
            CantidadPostes=((int(MetrosAux3)/25)+1)
            CantidadPostesSuspension=CantidadPostes*0.75
            CantidadPostesSuspension=round(CantidadPostesSuspension,0)
            CantidadPostesRetencion=-CantidadPostesSuspension+CantidadPostes
            Valor_conexion=Costos_LC['Poste 12m-1050kg retencion']*CantidadPostesRetencion+Costos_LC['Poste 12m-510kg suspension']*CantidadPostesSuspension+(int(MetrosAux3)/1000)*Costos_LC[varCLC]+(CantidadPostes/3)*Costos_LC['Puesta a tierra']
        
        Interconexion={
            "Transformadores": valorCostoTransformador,
            "Protecciones_MT": valorCostoProtecciones,
            "Medidores": valorCostoMedidores,
            "Linea_Conexion": Valor_conexion,
            }
        
        for clave in Equipos_Generacion:
            Equip_Gen_Total+=Equipos_Generacion[clave]
        for clave in BOS:
            BOS_Total+=BOS[clave]
        for clave in Interconexion:
            if clave!='Linea_Conexion':
                Total_Interconexion+=Interconexion[clave]
            
        Equip_Gen_Total=Equip_Gen_Total*C_InstaladaDC/1000
        BOS_Total=BOS_Total*C_InstaladaDC/1000
        Total_Interconexion=(Total_Interconexion*C_InstaladaDC/1000)+Valor_conexion
        Capex=Equip_Gen_Total+BOS_Total+Total_Interconexion+costoLegalizacion
        
        
        """Intereses durante la construcción"""
        
        InteresConstruccion=Capex*(valorPeriodoFinanciacion/12)*((valorTasaInteres/100)/2)
        
        """Calculo de R y el pago del prestamo"""
        porcenCostosDuros=valorPorcenCostosDuros
        plazoDeuda=valorPlazoDeuda
        interesDeuda=valorInteresDeuda
        tarifaPrestamista=valorTarifaPrestamista
        DCSRrequerido=valorDCSRminimo
        deudaAcapital=Capex*porcenCostosDuros/100
        interesDeudaList=[]
        capitalDeudaList=[]
        exponente=math.pow((1+interesDeuda/100),-plazoDeuda)
        Ani=(1-exponente)/(interesDeuda/100)
        R=deudaAcapital/Ani
        costoFinanAux=deudaAcapital
        valorSeguro=Capex*valorSeguroOyM/100
        AuxiliarLCOE=0
        
        for j in range(0,25):
            if j<plazoDeuda:
                interesDeudaList.append(costoFinanAux*interesDeuda/100)
                capitalDeudaList.append(R-interesDeudaList[j])
                costoFinanAux=costoFinanAux-capitalDeudaList[j]
            else:
                interesDeudaList.append(0)
                capitalDeudaList.append(0)
                
                
                
        """Ingresos"""
        Excedente=[]
        for i in range(0,29):
            if i<=23:
                degradacionEnergia=degradacionEnergia*(1-0.005)
                #print("año "+str(i+2)+str(" ")+str(degradacionEnergia))
                listaEnergia.append(degradacionEnergia*energiaPrimerAno)
                ENFICCkWlista.append(degradacionEnergia*energiaPrimerAno*ENFICC)
                perdidasIndis.append(degradacionEnergia*energiaPrimerAno*hIndisponibilidad/365/24)
                #TasaCTFE=TasaCTFE*(1+(tariffRateEscalated*tariffEscalationRate))
                TasaCTFE=TasaCTFE*(1+((valorPorcentajeIncrementoTarifa/100)*(valorTasaTarifaCostos/100)))
                tarifaEnergiaLista.append(valorTarifaEnergiaRed*TasaCTFE)
                tarifacostoPPAlista.append(costoPPA*TasaCTFE)
                CERElista.append(CERE*TasaCTFE)
                costoPPAAGlista.append(costoPPAAG*TasaCTFE)
                ElectrPagadaUsuario.append(valorTarifaEnergiaRed*TasaCTFE*demandaUsuarioSinAG)
                ElectrPagadaUsuarioAG.append(valorTarifaEnergiaRed*TasaCTFE*autoconsumoUsuario)
                AuxiliarLCOE+=listaEnergia[i]
                if calculoExactoVar==1:
                    Excedente.append(degradacionEnergia*valorEnergia[0])
            else:
                listaEnergia.append(0)
                Excedente.append(0)
        
        """Ingresos"""
        
        #Ingreso por venta de energía autoconsumo. 
        ingresoVentaAutAG=[]
        ingresoVentaAutAG.append(costoPPAAG*autoconsumoUsuario)
        #Ingreso por venta de excedentes a la red
        ingresoExcAG=[]
        ingresoExcAG.append(costoPPA*(energiaPrimerAno-energiaPrimerAno*hIndisponibilidad/365/24-autoconsumoUsuario))
        #Ingreso por cargo de confiabilidad
        ingresoCeret=[]
        ingresoCeret.append(energiaPrimerAno*ENFICC*CERE)
        #totalIngresos=[]
        #totalIngresos.append(ingresoVentaAutAG[0]+ingresoExcAG[0]+ingresoCeret[0]+interesDeuda[0])
        #print(Excedente)
        #print(len(Excedente))
        #print(len(listaEnergia))
        for j in range(1,len(tarifacostoPPAlista)):
            ingresoVentaAutAG.append(autoconsumoUsuario*costoPPAAGlista[j])
            if calculoExactoVar==1:
                ingresoExcAG.append((Excedente[j]-perdidasIndis[j])*tarifacostoPPAlista[j])
                #ingresoExcAG.append((listaEnergia[j]-perdidasIndis[j]-autoconsumoUsuario)*tarifacostoPPAlista[j])
            else:
                ingresoExcAG.append((listaEnergia[j]-perdidasIndis[j]-autoconsumoUsuario)*tarifacostoPPAlista[j])
            ingresoCeret.append(ENFICCkWlista[j]*CERElista[j])        
                
        #print(ingresoVentaAutAG)
        #print(perdidasIndis)        
        #print(ingresoExcAG)
        #print(ingresoCeret)
        """Egresos"""
        
        for i in range(0,25):
            if (valorRegalias>0):
                regaliasTotal.append((ingresoVentaAutAG[i]+ingresoExcAG[i]+ingresoCeret[i])*valorRegalias/100)
            else:
                regaliasTotal.append(0)
                if i<=(valorPeriodoInflacionOyM-1):
                    costosFijos.append((valorGastosOyM*Periodo*C_InstaladaDC)-valorAporteCliente*Periodo)
                    costosVariables.append((valorGastosVarOyM/100)*listaEnergia[i]*Periodo)
                    gerenciayAdministracion.append(administracion*Periodo)
                    arrTierra.append(valorArriendoTotal*Periodo)
                    seguroList.append(valorSeguro*Periodo)
                    totalEgresos.append(costosFijos[i]+costosVariables[i]+gerenciayAdministracion[i]+arrTierra[i]+seguroList[i]+regaliasTotal[i])
                    Periodo=Periodo*(1+valorInflacionOyM/100)
                else:
                    if i>(valorPeriodoInflacionOyM-1):
                        costosFijos.append((valorGastosOyM*Periodo*C_InstaladaDC)-valorAporteCliente*Periodo)
                        costosVariables.append((valorGastosVarOyM/100)*listaEnergia[i]*Periodo)
                        gerenciayAdministracion.append(administracion*Periodo)
                        arrTierra.append(valorArriendoTotal*Periodo)
                        seguroList.append(valorSeguro*Periodo)
                        totalEgresos.append(costosFijos[i]+costosVariables[i]+gerenciayAdministracion[i]+arrTierra[i]+seguroList[i]+regaliasTotal[i])
                        Periodo=Periodo*(1+valorInflacionOyM_dos/100)
        
        
        
        """Debito inicial cuenta de reserva"""
        
        nMesesCreserva=valorNumeroServicioDeuda
        cuotaInicialReserva=R*nMesesCreserva/12
        nMesesCreservaExpense=valorNumeroDeMesesGastoOyM
        inicialOM=(totalEgresos[0]/12)*nMesesCreservaExpense
        interesReserva=valorInteresCuentaReserva
        
        """Reservas y costo de financiamiento"""
        ReservaCostoFinan=0
        ReservaCostoFinan=valorTarifaPrestamista*valorPorcenCostosDuros*(Equip_Gen_Total+BOS_Total+Total_Interconexion+costoLegalizacion)-InteresConstruccion-cuotaInicialReserva-inicialOM
        
        """Costo total instalado"""
        CostoTotalInstalado=-1*ReservaCostoFinan+Capex
        CTotalMW=CostoTotalInstalado/C_InstaladaDC/1000
        Inversion=CostoTotalInstalado-deudaAcapital
        TIRAux=[]
        TIRAux.append(-Inversion)
        depreciacionAceleradaAux=CostoTotalInstalado
        TIRAux_dos=[]
        
        """De la ventana costos de capital"""
        
        reemplazoEquipo=valorReemplazoEquipo #Año
        costoReemplazo=valorCostoReemplazoEquipo #COP/Watt
        reemplazoEquipoDos=valorReemplazoEquipo_dos
        costoReemplazoDos=valorCostoReemplazoEquipo_dos 
        
        Acelerada=costoReemplazo*C_InstaladaDC*1000
        resReemplazoEquipos=[]
        balanceFinal=[]
        reservaDeuda=[]
        reservaDeuda.append(inicialOM+cuotaInicialReserva)
        interesDeuda=[]
        totalIngresos=[]
        EBITDA=[]
        EBITDAAcumulado=[]
        EBITDAux=0
        DSCR=[]
        ingresosSinInteres=[]
        ingresosAntesImpuestos=[]
        reservaCapitalTrabajo=[]
        reservaServicioDeuda=[]
        deduccionBaseGravable=[]
        acumuladoRenta=[]
        acumuladoRentaAux=0
        impuestoBaseGravable=[]
        impuestoRentaAcumulado=[]
        impuestoRentaAux=0
        ingresosDespuesImpuestos=[]
        TIRAux_dos.append(-Inversion)
        ajustePorReemplazo=[]
        VAN_dos=[]
        VAN_dos.append(-CostoTotalInstalado)
        
        for i in range(0,25):
            if i<reemplazoEquipo-1:
                resReemplazoEquipos.append((Acelerada)/((reemplazoEquipo-1)))
                balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                reservaDeuda.append(balanceFinal[i])
                interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                reservaCapitalTrabajo.append(0)
                reservaServicioDeuda.append(0)
                ajustePorReemplazo.append(0)
                #print(i)
            elif i==reemplazoEquipo-1:
                    resReemplazoEquipos.append(-Acelerada)
                    balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                    reservaDeuda.append(balanceFinal[i])
                    interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                    reservaCapitalTrabajo.append(0)
                    reservaServicioDeuda.append(0)
                    ajustePorReemplazo.append(Acelerada)
                    #print(i)
            elif i>reemplazoEquipo-1 and i<reemplazoEquipoDos-1:
                resReemplazoEquipos.append((vidaUtil)/(reemplazoEquipoDos-reemplazoEquipo-1))
                ajustePorReemplazo.append(0)
                if i==vidaUtil-1:
                    reservaCapitalTrabajo.append(inicialOM)
                    #print('no estoy entrando acá')
                else:
                    reservaCapitalTrabajo.append(0)
                if i==plazoDeuda:
                    reservaServicioDeuda.append(cuotaInicialReserva)
                else:
                    reservaServicioDeuda.append(0)
                if i<plazoDeuda:
                    balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                else:
                    balanceFinal.append(0)
                reservaDeuda.append(balanceFinal[i])
                interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                #print(i)
            else:
                if i<24:
                    reservaDeuda.append(0)
                resReemplazoEquipos.append(0)
                interesDeuda.append(0)
                #reservaCapitalTrabajo.append(0)
                reservaServicioDeuda.append(0)
                ajustePorReemplazo.append(0)
                if i==vidaUtil-1:
                    reservaCapitalTrabajo.append(inicialOM)
                else:
                    reservaCapitalTrabajo.append(0)
            totalIngresos.append(ingresoVentaAutAG[i]+ingresoExcAG[i]+ingresoCeret[i]+interesDeuda[i])
            EBITDA.append(totalIngresos[i]-totalEgresos[i])
            EBITDAux=EBITDAux+EBITDA[i]
            EBITDAAcumulado.append(EBITDAux)
            ingresosSinInteres.append(EBITDA[i]-interesDeudaList[i])
            ingresosAntesImpuestos.append(ingresosSinInteres[i]-capitalDeudaList[i]-resReemplazoEquipos[i]+reservaCapitalTrabajo[i]+reservaServicioDeuda[i]-ajustePorReemplazo[i])
            #print(len(reservaCapitalTrabajo[i]))
            TIRAux.append(ingresosAntesImpuestos[i])
            if i<plazoDeuda:
                DSCR.append((EBITDA[i]-resReemplazoEquipos[i])/R)
            else:
                DSCR.append("N/A")
            if depreciacion==0:
                """Depreciacion Lineal"""
                depreciacionAux=CostoTotalInstalado*valorTasaDepreciacionLn/100
                ValorDeRescate=0.1*CostoTotalInstalado
                if i<=valorTiempoDepreciacionLn-1:
                    depreciacionVal.append(depreciacionAux)
                    depreciacionAux2+=depreciacionAux
                    depreciacionlnAcumulada.append(depreciacionAux2)
                    valorLibro.append(CostoTotalInstalado+ValorDeRescate-depreciacionlnAcumulada[i])
                else:
                    depreciacionVal.append(0)
                    depreciacionlnAcumulada.append(0)
            elif depreciacion==1:
                """Depreciacion acelerada"""
                if i<=9:
                    montoDepreciar.append(depreciacionAceleradaAux)
                    depreciacionVal.append(depreciacionAceleradaAux*valorTasaDepreciacionAc/100)
                    depreciacionAcelAcumulada.append(depreciacionVal[i])
                    valorLibro.append(montoDepreciar[i]-depreciacionVal[i])
                    depreciacionAceleradaAux=valorLibro[i]
                    
                else:
                    montoDepreciar.append(0)
                    depreciacionVal.append(0)
                    depreciacionAcelAcumulada.append(0)
                    valorLibro.append(0)
            if impuestos==0:
                """El cliente declara renta"""
                if i==0:
                    deduccionBaseGravable.append(0)
                    acumuladoRenta.append(0)
                    RentaLiquida.append(ingresosSinInteres[i]-depreciacionVal[i])
                else:
                    if (acumuladoRenta[i-1]+RentaLiquida[i-1]*valorBaseGravable/100)<(CostoTotalInstalado*valorMaxDesCapex/100):
                        deduccionBaseGravable.append(RentaLiquida[i-1]*valorBaseGravable/100)
                        #print("Estoy en el primer if")
                    else:
                        if CostoTotalInstalado*valorMaxDesCapex/100-acumuladoRenta[i-1]<=0:
                            deduccionBaseGravable.append(0)
                        else:
                            deduccionBaseGravable.append(CostoTotalInstalado*valorMaxDesCapex/100-acumuladoRenta[i-1])
                            #print("Estoy en el tercer else")
                    acumuladoRentaAux=acumuladoRenta[i-1]
                    acumuladoRenta.append(deduccionBaseGravable[i]+acumuladoRentaAux)
                    RentaLiquida.append(ingresosSinInteres[i]-depreciacionVal[i]-deduccionBaseGravable[i])
                impuestoBaseGravable.append((RentaLiquida[i]-deduccionBaseGravable[i])*valorImpuestoRenta/100)
                impuestoRentaAux+=impuestoBaseGravable[i]
                impuestoRentaAcumulado.append(impuestoRentaAux)
            elif impuestos==1:
                """El cliente no declara renta"""
                RentaLiquida.append(0)
            ingresosDespuesImpuestos.append(ingresosAntesImpuestos[i]-impuestoBaseGravable[i])
            TIRAux_dos.append(ingresosDespuesImpuestos[i])
            VAN_dos.append(-totalEgresos[i])
                
        VANIngresos=npf.npv(valorEquityTIR/100,ingresosAntesImpuestos)
        AuxEgresosTotales=[]
        #VANEgresos=npf.npv(valorEquityTIR/100,totalEgresos)
        BC=VANIngresos/(Inversion)
        TIR=npf.irr(TIRAux)
        TIR_dos=npf.irr(TIRAux_dos)
        VAN=npf.npv(valorEquityTIR/100,TIRAux_dos)
        VAN_dosAux=npf.npv(valorEquityTIR/100,VAN_dos)
        valorUnitario=totalEgresos[0]/listaEnergia[0]
        LCOE=-VAN_dosAux/sum(listaEnergia)
        
        
        if contador>=1:
            reporteExcel(LCOE,C_InstaladaDC, energiaPrimerAno, CostoTotalInstalado, totalEgresos, valorUnitario, totalIngresos,
                             valorPorcenCostosDuros, Inversion, valorEquityTIR, plazoDeuda, impuestos, TIR,TIR_dos, VAN_dosAux,
                             VAN, DSCR, DCSRrequerido, valorDCSRpromedio, BC, impuestoBaseGravable, tarifacostoPPAlista,
                             ingresosDespuesImpuestos, interesDeudaList, contador, valorReporteEx, varLC, varCLC, track, valorGastosOyM,
                             valorAdministracionOyM, valorArriendoHa, valorAporteCliente, valorRegalias, valorPeriodoFinanciacion,
                             valorTasaInteres, valorPlazoDeuda, valorInteresDeuda, valorNumeroServicioDeuda, valorNumeroDeMesesGastoOyM,
                             valorInteresCuentaReserva, valorImpuestoRenta, depreciacion, valorTasaDepreciacionAc,
                             valorTiempoDepreciacionLn, valorTasaDepreciacionLn, valorTarifaEnergiaRed, valorCostoEnergiaPPA,
                             valorCostoEnergiaPPAgen, valorCERE, valorDuracionTarifaCostos, valorPorcentajeIncrementoTarifa,
                             valorTasaTarifaCostos, valorCambioEntrada)
        global data
        data.setdefault("Cambio hecho para comparacion", []).append(valorCambioEntrada)
        data.setdefault("Costo Nivelado De Energía",[]).append(LCOE)
        data.setdefault("Capacidad de energía",[]).append(C_InstaladaDC)
        data.setdefault("Costo Total Instalado",[]).append(CostoTotalInstalado)
        data.setdefault("Gastos Operativos",[]).append(totalEgresos[0])
        data.setdefault("Valor Unitario de costos operacionales",[]).append(valorUnitario)
        data.setdefault("Total Ingresos",[]).append(totalIngresos[0])
        data.setdefault("Inversión de capital",[]).append(Inversion)
        data.setdefault("TIR antes de impuestos",[]).append(TIR)
        data.setdefault("TIR después de impuestos",[]).append(TIR_dos)
        data.setdefault("Relación beneficio-costo",[]).append(BC)      
        return data
        
data={}

def conductorOptimo(energia, planta, gastos, costos, condicionales):
    
    Equip_Gen_Total=0
    BOS_Total=0
    Total_Interconexion=0
    TasaDolar=3965
    costoMWRTB=TasaDolar*50000
    costoLegalizacion=costoMWRTB*planta[0]
    CapacidadPlanta=planta[0]*1000*planta[1]*365*1.2
    CapacidadPlanta_dos=CapacidadPlanta/(planta[1]*365)
    variable=False
    iterador=0
    
    if condicionales[0]=='Parcialmente plano':
        valor_Terreno=50000000
    elif condicionales[0]=='Con pendiente superior al 10%':
        valor_Terreno=65000000
    elif condicionales[0]=='Con vegetación':
        valor_Terreno=75000000
    elif condicionales[0]=='No es uniforme presenta protuberancias':
        valor_Terreno=100000000
    else:
        valor_Terreno=50000000
    
    if condicionales[1]==1:
        Valor_trackers=850000000 
    else:
        Valor_trackers=0    
    
    Equipos_Generacion={"Paneles_Solares":costos[0], "Inversores":costos[1]}
    
    BOS={"Cableado_DC":costos[2],"Cableado_AC_BT":costos[3],"Sistema_Monitoreo":costos[4],"Trackers": 0,
         "Tuberias_conducciones": costos[5],"Cerramiento":costos[6],"SPT": costos[7],"Otros_materiales":costos[8],
         "Alquiler_maquinaria": costos[9],"Adecuacion_terrenos": 50000000,"Instalación": costos[10],
         "Imprevistos": costos[11],"Ingenieria_detalle": costos[12],"Estudio_suelos": costos[13],
         "Estudio_topograficos": costos[14],"Estudio_resistividad": costos[15],"Pruebas_campo": costos[16],
         "Campamento": costos[17],"SPDA": costos[18],
         }
    
    #Costos_Linea_Conexión
    Costos_LC={
        "Poste 12m-1050kg retencion": 7242621,
        "Poste 12m-510kg suspension":5509945,
        "ACSR 4 AWG": 18797225,
        "ACSR 2 AWG":21160588,
        "ACSR 1/0 AWG":22936110,
        "ACSR 2/0 AWG": 25388593,
        "ACSR 3/0 AWG": 28530033,
        "ACSR 4/0 AWG": 37373366,
        "ACSR 266 kcmil": 50239045,
        "ACSR 336 kcmil": 59884448,
        "ACSR 397 kcmil": 68948864,
        "ACSR 477 kcmil": 81187283,
        "ACSR 605 kcmil": 95211507,
        "ACSR 795 kcmil": 124576172,
        "Puesta a tierra": 462733
        }
    
    #Costo total línea conexión
    #varCLC=costos[19]
    distanciaDig=10
    while variable==False: 
        varCLC=calculoCalibre(planta[0],13.2,distanciaDig)
        CantidadPostes=((int(distanciaDig)/25)+1)
        CantidadPostesSuspension=CantidadPostes*0.75
        CantidadPostesSuspension=round(CantidadPostesSuspension,0)
        CantidadPostesRetencion=-CantidadPostesSuspension+CantidadPostes
        Valor_conexion=Costos_LC['Poste 12m-1050kg retencion']*CantidadPostesRetencion+Costos_LC['Poste 12m-510kg suspension']*CantidadPostesSuspension+(int(distanciaDig)/1000)*Costos_LC[varCLC]+(CantidadPostes/3)*Costos_LC['Puesta a tierra']
        
    
        Interconexion={
            "Transformadores": costos[20],
            "Protecciones_MT": costos[21],
            "Medidores": costos[22],
            "Linea_Conexion": Valor_conexion,
            }
        
        for clave in Equipos_Generacion:
            Equip_Gen_Total+=Equipos_Generacion[clave]
        for clave in BOS:
            BOS_Total+=BOS[clave]
        for clave in Interconexion:
            if clave!='Linea_Conexion':
                Total_Interconexion+=Interconexion[clave]
    
        
        Equip_Gen_Total=Equip_Gen_Total*CapacidadPlanta_dos/1000
        BOS_Total=BOS_Total*CapacidadPlanta_dos/1000
        Total_Interconexion=(Total_Interconexion*CapacidadPlanta_dos/1000)+Valor_conexion   
        Capex=Equip_Gen_Total+BOS_Total+Total_Interconexion+costoLegalizacion
        Equip_Gen_Total=0
        BOS_Total=0
        Total_Interconexion=0
        AuxiliarCosto=energia[1]
        AuxiliarCosto_dos=energia[3]
        AuxiliarCosto_tres=energia[4]
        Ingresos={}
        Egresos={}
        AutoconsumoList=[]
        ExcedenteList=[]
        AuxiliarEnergia=CapacidadPlanta
        Indisponibilidad=[]
        AuxiliarDegradacion=1
        EnficcList=[]
        valorGastosOyM=gastos[0]
        valorPeriodoInflacionOyM=gastos[1]
        valorAporteCliente=gastos[2]
        Periodo=1
        costosFijos=[]
        costosVariables=[]
        valorGastosVarOyM=gastos[3]
        listaEnergia=[]
        listaEnergia.append(CapacidadPlanta)
        gerenciayAdministracion=[]
        administracion=gastos[4]
        arrTierra=[]
        valorArriendoHa=gastos[5]
        Area_Proyecto=planta[0]*2
        valorArriendoTotal=valorArriendoHa*Area_Proyecto
        valorSeguro=Capex*gastos[6]/100
        seguroList=[]
        totalEgresos=[]
        valorInflacionOyM=gastos[7]
        valorInflacionOyM_dos=gastos[8]
        TirAux=[]
        totalIngresos=[]
        reemplazoEquipo=condicionales[3]
        resReemplazoEquipos=[]
        costoReemplazo=condicionales[4]
        Acelerada=costoReemplazo*CapacidadPlanta_dos*1000
        balanceFinal=[]
        reservaDeuda=[]
        nMesesCreservaExpense=condicionales[5]
        deudaAcapital=Capex*condicionales[2]/100
        interesDeudaList=[]
        capitalDeudaList=[]
        interesDeudaC=condicionales[6]
        plazoDeuda=condicionales[7]
        exponente=math.pow((1+interesDeudaC/100),-plazoDeuda)
        Ani=(1-exponente)/(interesDeudaC/100)
        R=deudaAcapital/Ani
        nMesesCreserva=condicionales[8]
        cuotaInicialReserva=R*nMesesCreserva/12
        interesReserva=condicionales[9]
        reservaCapitalTrabajo=[]
        reservaServicioDeuda=[]
        ajustePorReemplazo=[]
        reemplazoEquipoDos=condicionales[10]
        vidaUtil=25
        interesDeuda=[]
        EBITDA=[]
        valorTarifaPrestamista=condicionales[11]
        InteresConstruccion=Capex*(condicionales[13]/12)*((condicionales[12]/100)/2)
        valorEquityTIR=condicionales[14]
        VAN_dos=[]
        ingresosSinInteres=[]
        ingresosAntesImpuestos=[]
        costoFinanAux=deudaAcapital
    
        for i in range(0,25):
            AutoconsumoList.append(energia[0]*AuxiliarCosto)
            AuxiliarCosto=AuxiliarCosto*(1+(1*energia[2]/100))
            Indisponibilidad.append(AuxiliarEnergia*planta[3]/365/24)
            ExcedenteList.append((AuxiliarEnergia-energia[0]-Indisponibilidad[i])*AuxiliarCosto_dos)
            AuxiliarCosto_dos=AuxiliarCosto_dos*(1+(1*energia[2]/100))
            EnficcList.append((AuxiliarEnergia)*(planta[4]/100)*AuxiliarCosto_tres)
            AuxiliarCosto_tres=AuxiliarCosto_tres*(1+(1*energia[2]/100))
            AuxiliarDegradacion=AuxiliarDegradacion*(1-planta[2]/100)
            AuxiliarEnergia=CapacidadPlanta*AuxiliarDegradacion
            listaEnergia.append(AuxiliarEnergia)
                    
            if i<=(valorPeriodoInflacionOyM-1):
                costosFijos.append((valorGastosOyM*Periodo*CapacidadPlanta_dos)-valorAporteCliente*Periodo)
                costosVariables.append((valorGastosVarOyM/100)*listaEnergia[i]*Periodo)
                gerenciayAdministracion.append(administracion*Periodo)
                arrTierra.append(valorArriendoTotal*Periodo)
                seguroList.append(valorSeguro*Periodo)
                totalEgresos.append(costosFijos[i]+costosVariables[i]+gerenciayAdministracion[i]+arrTierra[i]+seguroList[i])
                Periodo=Periodo*(1+valorInflacionOyM/100)
            else:
                if i>(valorPeriodoInflacionOyM-1):
                    costosFijos.append((valorGastosOyM*Periodo*CapacidadPlanta_dos)-valorAporteCliente*Periodo)
                    costosVariables.append((valorGastosVarOyM/100)*listaEnergia[i]*Periodo)
                    gerenciayAdministracion.append(administracion*Periodo)
                    arrTierra.append(valorArriendoTotal*Periodo)
                    seguroList.append(valorSeguro*Periodo)
                    totalEgresos.append(costosFijos[i]+costosVariables[i]+gerenciayAdministracion[i]+arrTierra[i]+seguroList[i])
                    Periodo=Periodo*(1+valorInflacionOyM_dos/100)
            if i==0:
                inicialOM=(totalEgresos[0]/12)*nMesesCreservaExpense
                reservaDeuda.append(inicialOM+cuotaInicialReserva)
                ReservaCostoFinan=valorTarifaPrestamista*condicionales[2]*(Equip_Gen_Total+BOS_Total+Total_Interconexion+costoLegalizacion)-InteresConstruccion-cuotaInicialReserva-inicialOM
                CostoTotalInstalado=-1*ReservaCostoFinan+Capex
                CTotalMW=CostoTotalInstalado/CapacidadPlanta_dos/1000
                Inversion=CostoTotalInstalado-deudaAcapital
                TirAux.append(-Inversion)
                VAN_dos.append(-CostoTotalInstalado)
            if i<reemplazoEquipo-1:
                resReemplazoEquipos.append((Acelerada)/((reemplazoEquipo-1)))
                balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                reservaDeuda.append(balanceFinal[i])
                interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                reservaCapitalTrabajo.append(0)
                reservaServicioDeuda.append(0)
                ajustePorReemplazo.append(0)
                #print(i)
            elif i==reemplazoEquipo-1:
                    resReemplazoEquipos.append(-Acelerada)
                    balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                    reservaDeuda.append(balanceFinal[i])
                    interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                    reservaCapitalTrabajo.append(0)
                    reservaServicioDeuda.append(0)
                    ajustePorReemplazo.append(Acelerada)
                    #print(i)
            elif i>reemplazoEquipo-1 and i<reemplazoEquipoDos-1:
                resReemplazoEquipos.append((vidaUtil)/(reemplazoEquipoDos-reemplazoEquipo-1))
                ajustePorReemplazo.append(0)
                if i==vidaUtil-1:
                    reservaCapitalTrabajo.append(inicialOM)
                    #print('no estoy entrando acá')
                else:
                    reservaCapitalTrabajo.append(0)
                if i==plazoDeuda:
                    reservaServicioDeuda.append(cuotaInicialReserva)
                else:
                    reservaServicioDeuda.append(0)
                if i<plazoDeuda:
                    balanceFinal.append(reservaDeuda[i]+resReemplazoEquipos[i])
                else:
                    balanceFinal.append(0)
                reservaDeuda.append(balanceFinal[i])
                interesDeuda.append(((balanceFinal[i]+reservaDeuda[i])/2)*(interesReserva/100))
                #print(i)
            else:
                if i<24:
                    reservaDeuda.append(0)
                resReemplazoEquipos.append(0)
                interesDeuda.append(0)
                #reservaCapitalTrabajo.append(0)
                reservaServicioDeuda.append(0)
                ajustePorReemplazo.append(0)
                if i==vidaUtil-1:
                    reservaCapitalTrabajo.append(inicialOM)
                else:
                    reservaCapitalTrabajo.append(0)
            if i<plazoDeuda:
                interesDeudaList.append(costoFinanAux*interesDeudaC/100)
                capitalDeudaList.append(R-interesDeudaList[i])
                costoFinanAux=costoFinanAux-capitalDeudaList[i]
            else:
                interesDeudaList.append(0)
                capitalDeudaList.append(0)
            totalIngresos.append(AutoconsumoList[i]+ExcedenteList[i]+EnficcList[i]+interesDeuda[i])
            EBITDA.append(totalIngresos[i]-totalEgresos[i])
            TirAux.append(totalIngresos[i])
            VAN_dos.append(-totalEgresos[i])
            ingresosSinInteres.append(EBITDA[i]-interesDeudaList[i])
            ingresosAntesImpuestos.append(ingresosSinInteres[i]-capitalDeudaList[i]-resReemplazoEquipos[i]+reservaCapitalTrabajo[i]+reservaServicioDeuda[i]-ajustePorReemplazo[i])
            
        Ingresos['Autoconsumo']=AutoconsumoList
        Ingresos['Excedentes']=ExcedenteList
        Ingresos['Enficc']=EnficcList
        Ingresos['Interes Deuda']=interesDeuda
        Egresos['CostosFijos']=costosFijos
        #print(Capex)
        
        TIR=npf.irr(TirAux)*100
        VAN_dosAux=npf.npv(valorEquityTIR/100,VAN_dos)
        LCOE=-VAN_dosAux/sum(listaEnergia)
        VANIngresos=npf.npv(valorEquityTIR/100,ingresosAntesImpuestos)
        BC=VANIngresos/(Inversion)
            
        
        
        
        if TIR<13 or LCOE >220 or BC<1.3:
            variable=True
            print("la distancia crítica de viabilidad del proyecto es: "+str(distanciaDig)+" metros, el proceso se realizó en: "+str(iterador)+" iteraciones"+str(" .El calibre de la línea es ")+str(varCLC))
            print("La tir es: "+str(TIR))
            print("El LCOE es: "+str(LCOE))
            print("La relación beneficio costo es: "+str(BC))
            #cadena="la distancia crítica de viabilidad del proyecto es: "+str(distanciaDig)+" metros, el proceso se realizó en: "+str(iterador)+" iteraciones"+str(" .El calibre de la línea es ")+str(varCLC)"\n "
        else:
            distanciaDig+=1
            iterador+=1
            
               
            
        
    
    
    
    return "Holi"